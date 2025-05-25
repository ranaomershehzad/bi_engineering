import openpyxl as xl

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Give column 4 a name
    sheet["D1"] = "Discounted Price"

    for row in range(2,sheet.max_row + 1):
        # Read the price column
        price_cell = sheet.cell(row,3)

        # Apply the discount on price values per row
        discounted_value = price_cell.value * 0.9

        # Give it a place to write
        discounted_cell = sheet.cell(row,4)

        # Assign the values to write onto
        discounted_cell.value = discounted_value

    wb.save(filename)

process_workbook("transactions_new.xlsx")